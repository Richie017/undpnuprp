from openpyxl.reader.excel import load_workbook

__author__ = 'Mahmud'


class GenericImporter(object):
    def __init__(self, **kwargs):
        super().__init__()

    @classmethod
    def import_from_excel(cls, model=None, filename=None, user=None, importer_config=None, request=None):
        imported_items = []
        wb = load_workbook(filename=filename, read_only=True)
        all_ws = wb.get_sheet_names()
        for ws_name in all_ws:
            ws = wb.get_sheet_by_name(ws_name)
            importing_columns = importer_config.columns.all()
            importing_columns = sorted(importing_columns, key=lambda x: x.column)
            row_index = -1
            enable_import = False
            for row in ws.iter_rows():
                row_index += 1
                column_index = -1
                ic_index = 0
                value_dict = dict()
                if not enable_import:
                    enable_import = True if row_index >= importer_config.starting_row else False
                if not enable_import:
                    continue

                for cell in row:
                    column_index += 1
                    if ic_index >= len(importing_columns):
                        break

                    column_config = importing_columns[ic_index]
                    if column_config.column > column_index:
                        continue

                    if column_config.column == column_index and column_config.ignore:
                        ic_index += 1
                        continue

                    if column_config.column == column_index:
                        ic_index += 1
                        value_type = type(cell.value)
                        stripped_value = cell.value.strip() if value_type == str else cell.value
                        value_dict[str(column_config.column)] = stripped_value

                if ic_index < len(importing_columns) - 1:
                    raise ValueError('Not all required data are provided')
                imported_items.append(
                    model.import_item(importer_config, importing_columns, value_dict, user=user, request=request))
        return imported_items
