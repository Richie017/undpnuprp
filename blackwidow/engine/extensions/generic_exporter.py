from openpyxl import Workbook


class GenericExporter(object):
    def __init__(self, header, elements, query_set, filename, title):
        self.header = header
        self.elements = elements
        self.query_set = query_set
        self.filename = filename
        self.title = title


    def export_to_exel(self):

        wb = Workbook()
        ws = wb.active
        dest_filename = self.filename
        ws.title = self.title

        for col_idx, head in enumerate(self.header):
            # col=get_column_letter(col_idx)
            ws.cell(row=0, column=col_idx).value = str(head)

        for row_idx, data in enumerate(self.query_set):
            for col_idx in range(len(self.header)):
                if self.elements[col_idx]:
                    if self.elements[col_idx] == "date_created":
                        l_date = str(getattr(data, self.elements[col_idx]).date()).replace("-", "/").split('/')
                        formated_date = '/'.join([l_date[i] for i in [2, 1, 0]])
                        ws.cell(row=row_idx + 1, column=col_idx).value = formated_date
                    else:
                        ws.cell(row=row_idx + 1, column=col_idx).value = getattr(data, self.elements[col_idx])

        for id, column in enumerate(ws.column_dimensions):
            ws.column_dimensions[column].width = len(self.header[id]) + 1

        wb.save(filename=dest_filename)









