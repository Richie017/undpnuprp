import os
import uuid
from decimal import Decimal

from django.conf import settings
from django.db import models
from django.db.models import Sum
from openpyxl.reader.excel import load_workbook

from blackwidow.core.models import FileQueueEnum
from blackwidow.core.models.contracts.organizationdomainentity import OrganizationDomainEntity
from blackwidow.engine.extensions import Clock
from blackwidow.engine.file_handlers.file_path_handler import FilePathHandler
from blackwidow.engine.routers.database_router import BWDatabaseRouter

S3_STATIC_ENABLED = settings.S3_STATIC_ENABLED

__author__ = "Ziaul Haque"

COLUMN_IDENTIFIER_DICT = {
    2020: 16,  # "P"
    2021: 18,  # "R"
    2022: 20,  # "T"
    2023: 22,  # "V"
    2024: 24,  # "X"
}

ROW_IDENTIFIER_LIST = [
    40, 41, 42, 43,
    46, 47, 48,
    51, 52, 53, 54, 55,
    59, 60, 61,
    64, 65, 66,
    69, 70, 71,
    75, 76,
    80, 81, 82, 83,
    86, 87, 88,
    92, 93, 94,
    98, 99, 100, 101,
    104, 105,
    109, 110,
    115, 116,
    119, 120, 121, 122, 123, 124,
    126, 127, 128
]
CUMULATIVE_COLUMN = 28
DEFAULT_ACHIEVEMENT_COLUMN = 15


class PMFUploadedFileQueue(OrganizationDomainEntity):
    pmf_report = models.ForeignKey('approvals.PMFReport')
    status = models.CharField(max_length=255, default=FileQueueEnum.SCHEDULED)
    file = models.ForeignKey('core.FileObject')

    class Meta:
        app_label = 'approvals'

    @property
    def render_code(self):
        return self.code

    @classmethod
    def get_manage_buttons(cls):
        return []

    @classmethod
    def get_object_inline_buttons(cls):
        return []

    @classmethod
    def process_scheduled_queues(cls):
        queue_queryset = cls.objects.filter(
            status=FileQueueEnum.SCHEDULED,
            pmf_report__city__isnull=False
        ).prefetch_related('pmf_report').order_by('pk')

        temporary_files = []
        for queue_instance in queue_queryset:
            # update queue status (set to processing)
            queue_instance.status = FileQueueEnum.PROCESSING
            queue_instance.save()

            filename = FilePathHandler.get_absolute_path(queue_instance.file.file)
            if S3_STATIC_ENABLED:
                temporary_files.append(filename)

            print("Loading file {0}".format(filename))
            city = queue_instance.pmf_report.city
            year = queue_instance.pmf_report.year
            month = queue_instance.pmf_report.month
            wb = load_workbook(filename=filename, read_only=True)
            cls.import_from_workbook(city=city, year=year, month=month, workbook=wb, queue=queue_instance)

        # after successfully import, remove local file
        for temp_file in temporary_files:
            try:
                os.remove(temp_file)
            except:
                pass

    @classmethod
    def import_from_workbook(cls, city, year, month, workbook, queue):
        from undp_nuprp.approvals.models import CityWisePMFReportAchievement
        ws = workbook.active
        achievement_queryset = CityWisePMFReportAchievement.objects.filter(city=city, year=year, month=month)
        achievement_dict = {}
        for achievement_instance in achievement_queryset:
            _key = (achievement_instance.row, achievement_instance.column)
            achievement_dict[_key] = achievement_instance

        column = COLUMN_IDENTIFIER_DICT.get(year, DEFAULT_ACHIEVEMENT_COLUMN)
        creatable_entries = list()
        updatable_entries = list()

        time_now = Clock.timestamp()
        for row in ROW_IDENTIFIER_LIST:
            cell_value = ws.cell(row=row, column=column).value
            if cell_value:
                try:
                    cell_value = Decimal(cell_value)
                    _key = (row, column)
                    if _key in achievement_dict.keys():
                        instance = achievement_dict.get(_key)
                        instance.last_updated = time_now
                        time_now += 1

                        updatable_entries.append(instance)
                    else:
                        instance = CityWisePMFReportAchievement()
                        instance.type = CityWisePMFReportAchievement.__name__
                        instance.organization = city.organization
                        instance.city = city
                        instance.year = year
                        instance.month = month
                        instance.row = row
                        instance.column = column
                        instance.date_created = time_now
                        instance.last_updated = time_now
                        time_now += 1

                        creatable_entries.append(instance)

                    instance.tsync_id = uuid.uuid4() if instance.tsync_id is None else instance.tsync_id
                    instance.achieved = cell_value
                except:
                    pass

        if len(creatable_entries) > 0:
            CityWisePMFReportAchievement.objects.bulk_create(creatable_entries)

        if len(updatable_entries) > 0:
            CityWisePMFReportAchievement.objects.bulk_update(updatable_entries)

        # update queue status (set to finished)
        queue.status = FileQueueEnum.COMPLETED
        queue.save()

    @classmethod
    def export_items(cls, city, year, worksheet):
        from undp_nuprp.approvals.models import CityWisePMFReportAchievement
        if city:
            achievement_queryset = CityWisePMFReportAchievement.objects.using(
                BWDatabaseRouter.get_export_database_name()
            ).filter(
                city=city
            ).values('row').annotate(Sum('achieved'))
            for item in achievement_queryset:
                row, column = item['row'], CUMULATIVE_COLUMN
                worksheet.cell(row=row, column=column).value = item['achieved__sum']
        else:
            achievement_queryset = CityWisePMFReportAchievement.objects.using(
                BWDatabaseRouter.get_export_database_name()
            ).values('row').annotate(Sum('achieved'))
            for item in achievement_queryset:
                row, column = item['row'], CUMULATIVE_COLUMN
                worksheet.cell(row=row, column=column).value = item['achieved__sum']
            achievement_queryset = CityWisePMFReportAchievement.objects.using(
                BWDatabaseRouter.get_export_database_name()
            ).values('year', 'row').annotate(Sum('achieved'))
            for item in achievement_queryset:
                row, column = item['row'], COLUMN_IDENTIFIER_DICT.get(item['year'], DEFAULT_ACHIEVEMENT_COLUMN)
                worksheet.cell(row=row, column=column).value = item['achieved__sum']

        return worksheet
