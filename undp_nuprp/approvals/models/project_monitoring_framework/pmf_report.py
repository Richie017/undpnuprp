import os
from datetime import datetime

from django import forms
from django.conf import settings
from django.db import models
from openpyxl import load_workbook, Workbook

from blackwidow.core.mixins.fieldmixin import GenericModelChoiceField
from blackwidow.core.models import Geography, ExportFileObject, ErrorLog
from blackwidow.core.models.contracts.organizationdomainentity import OrganizationDomainEntity
from blackwidow.engine.decorators.route_partial_routes import route
from blackwidow.engine.decorators.utility import is_object_context, decorate
from blackwidow.engine.enums.modules_enum import ModuleEnum
from blackwidow.engine.enums.view_action_enum import ViewActionEnum
from blackwidow.engine.file_handlers.aws_file_writer import AWSFileWriter
from blackwidow.engine.file_handlers.file_path_handler import FilePathHandler
from blackwidow.engine.routers.database_router import BWDatabaseRouter
from config.aws_s3_config import MEDIA_DIRECTORY

EXPORT_FILE_ROOT = settings.EXPORT_FILE_ROOT
S3_STATIC_ENABLED = settings.S3_STATIC_ENABLED

__author__ = "Ziaul Haque"


@decorate(is_object_context,
          route(route='monthly-pmf-report', group='Project Monitoring Framework', module=ModuleEnum.Analysis,
                display_name='Monthly Report', group_order=6, item_order=2))
class PMFReport(OrganizationDomainEntity):
    month = models.IntegerField(default=1)
    year = models.IntegerField(default=2020)
    city = models.ForeignKey('core.Geography', null=True, blank=True)

    class Meta:
        app_label = 'approvals'

    @classmethod
    def default_order_by(cls):
        return ['-year', '-month']

    @property
    def code_prefix(self):
        return "PMF"

    @property
    def render_month(self):
        return datetime.now().replace(year=self.year, month=self.month).strftime("%B, %Y")

    @property
    def render_city(self):
        return self.city if self.city else "All Cities"

    @property
    def render_uploaded_excel(self):
        from undp_nuprp.approvals.models import PMFUploadedFileQueue
        uploaded_excel = PMFUploadedFileQueue.objects.filter(pmf_report=self).order_by('-pk').first()
        return uploaded_excel.file if uploaded_excel else "N/A"

    @classmethod
    def details_view_fields(cls):
        return [
            'render_code', 'render_city', 'render_month', 'render_uploaded_excel',
            'date_created:Created On', 'last_updated', 'created_by', 'last_updated_by'
        ]

    @classmethod
    def table_columns(cls):
        return ['render_code', 'render_city', 'render_month', 'last_updated', 'last_updated_by']

    @classmethod
    def get_button_title(cls, button=ViewActionEnum.Details):
        if button == ViewActionEnum.AdvancedExport:
            return "Export"
        return button

    @classmethod
    def get_manage_buttons(cls):
        return [ViewActionEnum.Create, ViewActionEnum.Edit, ViewActionEnum.AdvancedExport, ]

    @classmethod
    def get_object_inline_buttons(cls):
        return [ViewActionEnum.Details, ViewActionEnum.Edit, ]

    def details_link_config(self, **kwargs):
        return [
            dict(
                name='Edit',
                action='edit',
                icon='fbx-rightnav-edit',
                ajax='0',
                url_name=self.__class__.get_route_name(action=ViewActionEnum.Edit)
            ),
        ]

    @classmethod
    def get_export_dependant_fields(cls):

        class AdvancedExportDependentForm(forms.Form):
            def __init__(self, data=None, files=None, instance=None, prefix='', **kwargs):
                super(AdvancedExportDependentForm, self).__init__(data=data, files=files, prefix=prefix, **kwargs)

                self.fields['city'] = \
                    GenericModelChoiceField(
                        queryset=Geography.objects.using(BWDatabaseRouter.get_export_database_name()).filter(
                            level__name='Pourashava/City Corporation',
                        ),
                        label='Select City',
                        empty_label='All Cities',
                        required=False,
                        widget=forms.Select(
                            attrs={
                                'class': 'select2 kpi-filter',
                                'width': '220',
                                'data-kpi-filter-role': 'city'
                            }
                        )
                    )

            class Meta:
                fields = ('city',)

        return AdvancedExportDependentForm

    @classmethod
    def export_to_excel(cls, organization, user, filename, query_params, **kwargs):
        from undp_nuprp.approvals.models import PMFUploadedFileQueue
        path = os.path.join(EXPORT_FILE_ROOT)
        if not os.path.exists(path):
            os.makedirs(path)

        generated_file_path = path + os.sep + str(filename) + '.xlsx'

        _city = query_params.get('city')
        if _city:
            queue_instance = PMFUploadedFileQueue.objects.using(
                BWDatabaseRouter.get_export_database_name()
            ).filter(pmf_report__city=_city).order_by(
                '-pmf_report__year', '-pmf_report__month', '-pk').first()
        else:
            queue_instance = PMFUploadedFileQueue.objects.using(
                BWDatabaseRouter.get_export_database_name()
            ).filter(pmf_report__city__isnull=True).order_by(
                '-pmf_report__year', '-pmf_report__month', '-pk').first()

        temporary_files = []
        if queue_instance:
            template_path = FilePathHandler.get_absolute_path(queue_instance.file.file)
            if S3_STATIC_ENABLED:
                temporary_files.append(template_path)

            print("Loading file {0}".format(template_path))

            workbook = load_workbook(filename=template_path)
            worksheet = workbook.active
            worksheet = PMFUploadedFileQueue.export_items(city=_city, year=2020, worksheet=worksheet)
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Project Monitoring Framework"

        workbook.save(filename=generated_file_path)

        # Uploading the exported file to AMAZON S3
        if S3_STATIC_ENABLED:
            from config.aws_s3_config import STATIC_EXPORT_MEDIA_DIRECTORY
            try:
                with open(generated_file_path, 'rb') as content_file:
                    content = content_file.read()
                    s3_file_name = STATIC_EXPORT_MEDIA_DIRECTORY + "/" + str(filename) + '.xlsx'
                    generated_file_path = s3_file_name.replace(MEDIA_DIRECTORY, '', 1)
                    AWSFileWriter.upload_file_with_content(file_name=s3_file_name, content=content)
                temporary_files.append(generated_file_path)

                # after successfully upload to AWS S3, remove local file
                for temp_file in temporary_files:
                    try:
                        os.remove(temp_file)
                    except:
                        pass
            except Exception as exp:
                ErrorLog.log(exp=exp)

        export_file_object = ExportFileObject()
        export_file_object.path = generated_file_path
        export_file_object.name = filename
        export_file_object.file = generated_file_path
        export_file_object.extension = '.xlsx'
        export_file_object.organization = organization
        export_file_object.save()

        return filename, generated_file_path
