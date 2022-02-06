"""
Created by tareq on 2/15/18
"""
from datetime import date

from django import forms
from django.db.models.expressions import F
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font

from blackwidow.core.models.config.exporter_column_config import ExporterColumnConfig
from blackwidow.core.models.config.exporter_config import ExporterConfig
from blackwidow.engine.decorators.route_partial_routes import route
from blackwidow.engine.decorators.utility import is_object_context, decorate
from blackwidow.engine.enums.modules_enum import ModuleEnum
from blackwidow.engine.enums.view_action_enum import ViewActionEnum
from blackwidow.engine.extensions.bw_titleize import bw_titleize
from blackwidow.engine.routers.database_router import BWDatabaseRouter
from undp_nuprp.approvals.models.savings_and_credits.base.cdc_monthly_report_field import CDCMonthlyReportField
from undp_nuprp.approvals.models.savings_and_credits.cdc_reports.cdc_monthly_report import CDCMonthlyReport

__author__ = 'Tareq', 'Shuvro'


@decorate(is_object_context,
          route(route='approved-cdc-monthly-report', group='Savings & Credit Reports', module=ModuleEnum.Execute,
                display_name='Approved CDC Reports', group_order=1, item_order=4))
class ApprovedCDCMonthlyReport(CDCMonthlyReport):
    class Meta:
        app_label = 'approvals'
        proxy = True

    @classmethod
    def distinct_fields(cls):
        return ['parent_id']

    @classmethod
    def get_manage_buttons(cls):
        return [ViewActionEnum.AdvancedExport]

    @classmethod
    def get_button_title(cls, button=ViewActionEnum.Details):
        if button == ViewActionEnum.AdvancedExport:
            return "Export"

    def details_link_config(self, **kwargs):
        return []

    @classmethod
    def get_export_dependant_fields(cls):
        class AdvancedExportDependentForm(forms.Form):
            def __init__(self, data=None, files=None, instance=None, prefix='', **kwargs):
                super(AdvancedExportDependentForm, self).__init__(data=data, files=files, prefix=prefix, **kwargs)
                today = date.today()
                year_choices = tuple()
                for y in range(2000, 2100):
                    year_choices += ((y, str(y)),)

                self.fields['report_type'] = forms.CharField(
                    label='Type of Report',
                    required=True,
                    widget=forms.Select(
                        attrs={'class': 'select2'},
                        choices=(
                            ('baseline', 'Baseline'),
                            ('monthly', 'Monthly'),
                        )
                    )
                )
                self.fields['year'] = forms.ChoiceField(
                    choices=year_choices,
                    widget=forms.Select(
                        attrs={'class': 'hidden-select2', 'width': '220'},
                    ),

                    initial=today.year
                )

                self.fields['year'].widget = forms.HiddenInput()

        class Meta:
            fields = ('report_type', 'year')

        return AdvancedExportDependentForm

    @classmethod
    def exporter_config(cls, organization=None, **kwargs):
        ExporterConfig.objects.filter(model=cls.__name__, organization=organization).delete()
        exporter_config, created = ExporterConfig.objects.get_or_create(model=cls.__name__, organization=organization)
        cdc_monthly_report_fields = CDCMonthlyReportField.objects.using(
            BWDatabaseRouter.get_export_database_name()
        ).values_list('name', flat=True).order_by('weight')

        columns = [
            ExporterColumnConfig(column=0, column_name='Division', property_name='division', ignore=False),
            ExporterColumnConfig(column=1, column_name='City', property_name='city', ignore=False),
            ExporterColumnConfig(column=2, column_name='Ward', property_name='ward', ignore=False),
            ExporterColumnConfig(column=3, column_name='Cluster Name', property_name='cluster_name', ignore=False),
            ExporterColumnConfig(column=4, column_name='Cluster Id', property_name='cluster_id', ignore=False),
            ExporterColumnConfig(column=5, column_name='CDC Report ID', property_name='cdc_report_id', ignore=False),
            ExporterColumnConfig(column=6, column_name='CDC Name', property_name='cdc_name', ignore=False),
            ExporterColumnConfig(column=7, column_name='CDC ID', property_name='cdc_id', ignore=False),
            ExporterColumnConfig(column=8, column_name='Month', property_name='cdc_report_month', ignore=False),
            ExporterColumnConfig(column=9, column_name='Year', property_name='cdc_report_year', ignore=False),
            ExporterColumnConfig(column=10, column_name='User Contact Number', property_name='phone_number',
                                 ignore=False),
            ExporterColumnConfig(column=11, column_name='User', property_name='user_name', ignore=False)
        ]
        column_no = 12

        for cdc_field_name in cdc_monthly_report_fields:
            columns.append(
                ExporterColumnConfig(column=column_no, column_name=cdc_field_name, property_name='details_field',
                                     ignore=False))
            column_no += 1

        columns.append(ExporterColumnConfig(column=column_no, column_name='Remarks', property_name='cdc_report_remarks',
                                            ignore=False))

        for c in columns:
            c.save(organization=organization, **kwargs)
            exporter_config.columns.add(c)
        return exporter_config

    def export_item(self, workbook=None, columns=None, row_number=None, **kwargs):
        return self.pk, row_number + 1

    @classmethod
    def finalize_export(cls, workbook=None, row_number=None, query_set=None, **kwargs):
        return workbook

    @classmethod
    def initialize_export(cls, workbook=None, columns=None, row_number=None, query_set=None, **kwargs):
        _today = date.today()
        query_params = kwargs.get('query_params')
        report_type = query_params.get('report_type', 'baseline')
        report_year = query_params.get('year', _today.year)

        approved_cdc_report_ids = ApprovedCDCMonthlyReport.objects.using(
            BWDatabaseRouter.get_export_database_name()
        ).distinct('parent_id').values_list('pk', flat=True).order_by('parent_id', '-on_spot_creation_time')

        if report_type == 'baseline':
            cdc_report_export_queryset = ApprovedCDCMonthlyReport.objects.using(
                BWDatabaseRouter.get_export_database_name()
            ).filter(
                is_baseline=True,
                pk__in=list(approved_cdc_report_ids)
            ).order_by('-on_spot_creation_time')
        else:
            cdc_report_export_queryset = ApprovedCDCMonthlyReport.objects.using(
                BWDatabaseRouter.get_export_database_name()
            ).filter(
                is_baseline=False,
                pk__in=list(approved_cdc_report_ids)
            ).order_by('-on_spot_creation_time')

        cdc_report_export_values = cdc_report_export_queryset.annotate(
            cdc_report_id=F('pk'),
            cdc_report_month=F('month'),
            cdc_report_year=F('year'),
            cdc_report_remarks=F('remarks'),
            # scg_report_location = Concat(F('location__latitude', output_field=models.TextField()), Value(' '), F('location__longitude', output_field=models.TextField())),
            cdc_id=F('cdc__assigned_code'),
            cdc_name=F('cdc__name'),
            cluster_id=F('cdc__parent__assigned_code'),
            cluster_name=F('cdc__parent__name'),
            ward=F('cdc__address__geography__name'),
            city=F('cdc__address__geography__parent__name'),
            division=F('cdc__address__geography__parent__parent__name'),
            phone_number=F('created_by__phones__phone'),
            user_name=F('created_by__name')
        ).values(
            'cdc_report_id', 'cdc_report_month', 'cdc_report_year', 'cdc_report_remarks', 'cdc_id', 'cdc_name',
            'cluster_id', 'cluster_name', 'ward', 'city', 'division', 'phone_number', 'user_name')

        cdc_report_extra_values = cdc_report_export_queryset.annotate(
            cdc_report_id=F('pk'),
            field_name=F('field_values__field__name'),
            field_value=F('field_values__value')).values('cdc_report_id', 'field_name', 'field_value')

        pk_wise_cdc_extra_dict = dict()
        for cdc_report_extra in cdc_report_extra_values:
            _cdc_report_id = cdc_report_extra['cdc_report_id']
            _field = cdc_report_extra['field_name'].lower().strip(' ')
            _value = cdc_report_extra['field_value']
            if _cdc_report_id not in pk_wise_cdc_extra_dict:
                pk_wise_cdc_extra_dict[_cdc_report_id] = dict()
            pk_wise_cdc_extra_dict[_cdc_report_id][_field] = _value

        row_index = row_number + 1
        for column in columns:
            workbook.cell(row=row_index, column=column.column + 1).value = bw_titleize(column.column_name)
            workbook.cell(row=row_index, column=column.column + 1).font = Font(bold=True)
            workbook.cell(row=row_number, column=column.column + 1).alignment = Alignment(wrap_text=True)

        row_index += 1
        for cdc_report_export_info in cdc_report_export_values:
            _cdc_report_id = cdc_report_export_info['cdc_report_id']
            for column in columns:
                column_index = column.column + 1
                column_name = column.column_name.lower().strip(' ')
                _value = cdc_report_export_info[column.property_name] if column.property_name != 'details_field' else \
                    pk_wise_cdc_extra_dict[_cdc_report_id][column_name] if column_name in pk_wise_cdc_extra_dict[
                        _cdc_report_id] else ''
                workbook.cell(row=row_index, column=column_index).value = str(_value) if _value else ''
            row_index += 1

        return workbook, row_index
