"""
Created by tareq on 2/15/18
"""
import re
from datetime import date

from django import forms
from django.db.models.expressions import F
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font

from blackwidow.core.models.config.exporter_column_config import ExporterColumnConfig
from blackwidow.core.models.config.exporter_config import ExporterConfig
from blackwidow.engine.decorators.route_partial_routes import route
from blackwidow.engine.decorators.utility import is_object_context, decorate
from blackwidow.engine.enums.modules_enum import ModuleEnum
from blackwidow.engine.enums.reachout_level_enum import ReachoutLevelEnum
from blackwidow.engine.enums.view_action_enum import ViewActionEnum
from blackwidow.engine.extensions.bw_titleize import bw_titleize
from blackwidow.engine.routers.database_router import BWDatabaseRouter
from undp_nuprp.approvals.models.savings_and_credits.base.scg_monthly_report_field import SCGMonthlyReportField
from undp_nuprp.approvals.models.savings_and_credits.scg_reports.scg_monthly_report import SCGMonthlyReport

__author__ = 'Tareq'


@decorate(is_object_context,
          route(route='approved-scg-monthly-report', group='Savings & Credit Reports', module=ModuleEnum.Execute,
                display_name='Approved SCG Reports', group_order=1, item_order=2))
class ApprovedSCGMonthlyReport(SCGMonthlyReport):
    class Meta:
        app_label = 'approvals'
        proxy = True

    @classmethod
    def distinct_fields(cls):
        return ['parent_id']

    @classmethod
    def get_manage_buttons(cls):
        return [ViewActionEnum.AdvancedExport]

    @classmethod
    def get_button_title(cls, button=ViewActionEnum.Details):
        if button == ViewActionEnum.AdvancedExport:
            return "Export"

    def details_link_config(self, **kwargs):
        return []

    @classmethod
    def get_export_dependant_fields(cls):

        class AdvancedExportDependentForm(forms.Form):
            def __init__(self, data=None, files=None, instance=None, prefix='', **kwargs):
                super(AdvancedExportDependentForm, self).__init__(data=data, files=files, prefix=prefix, **kwargs)
                today = date.today()
                year_choices = tuple()
                for y in range(2000, 2100):
                    year_choices += ((y, str(y)),)

                self.fields['report_type'] = forms.CharField(
                    label='Type of Report',
                    required=True,
                    widget=forms.Select(
                        attrs={'class': 'select2'},
                        choices=(
                            ('baseline', 'Baseline'),
                            ('monthly', 'Monthly'),
                        )
                    )
                )
                self.fields['year'] = forms.ChoiceField(
                    choices=year_choices,
                    widget=forms.Select(
                        attrs={'class': 'hidden-select2', 'width': '220'},
                    ),

                    initial=today.year
                )

                self.fields['year'].widget = forms.HiddenInput()

        class Meta:
            fields = ('report_type', 'year')

        return AdvancedExportDependentForm

    @classmethod
    def exporter_config(cls, organization=None, **kwargs):
        ExporterConfig.objects.filter(model=cls.__name__, organization=organization).delete()
        exporter_config, created = ExporterConfig.objects.get_or_create(model=cls.__name__, organization=organization)
        scg_monthly_report_fields = SCGMonthlyReportField.objects.using(
            BWDatabaseRouter.get_export_database_name()
        ).values_list('name', flat=True).order_by('weight')

        columns = [
            ExporterColumnConfig(column=0, column_name='Division', property_name='division', ignore=False),
            ExporterColumnConfig(column=1, column_name='City', property_name='city', ignore=False),
            ExporterColumnConfig(column=2, column_name='Ward', property_name='ward', ignore=False),
            ExporterColumnConfig(column=3, column_name='Cluster Name', property_name='cluster_name', ignore=False),
            ExporterColumnConfig(column=4, column_name='Cluster Id', property_name='cluster_id', ignore=False),
            ExporterColumnConfig(column=5, column_name='CDC Name', property_name='cdc_name', ignore=False),
            ExporterColumnConfig(column=6, column_name='CDC ID', property_name='cdc_id', ignore=False),
            ExporterColumnConfig(column=7, column_name='Primary Group Name', property_name='primary_group_name',
                                 ignore=False),
            ExporterColumnConfig(column=8, column_name='Primary Group ID', property_name='primary_group_id',
                                 ignore=False),
            ExporterColumnConfig(column=9, column_name='SCG Name', property_name='scg_name', ignore=False),
            ExporterColumnConfig(column=10, column_name='SCG Report ID', property_name='scg_report_id', ignore=False),
            ExporterColumnConfig(column=11, column_name='Month', property_name='scg_report_month', ignore=False),
            ExporterColumnConfig(column=12, column_name='Year', property_name='scg_report_year', ignore=False),
            ExporterColumnConfig(column=13, column_name='User Contact Number', property_name='phone_number',
                                 ignore=False),
            ExporterColumnConfig(column=14, column_name='User', property_name='user_name', ignore=False)
        ]

        column_no = 15

        for cdc_field_name in scg_monthly_report_fields:
            columns.append(
                ExporterColumnConfig(column=column_no, column_name=cdc_field_name, property_name='details_field',
                                     ignore=False))
            column_no += 1

        columns.append(
            ExporterColumnConfig(column=column_no, column_name='Remarks', property_name='scg_report_remarks',
                                 ignore=False))

        for c in columns:
            c.save(organization=organization, **kwargs)
            exporter_config.columns.add(c)
        return exporter_config

    def export_item(self, workbook=None, columns=None, row_number=None, **kwargs):
        return self.pk, row_number + 1

    @classmethod
    def finalize_export(cls, workbook=None, row_number=None, query_set=None, **kwargs):
        return workbook

    @classmethod
    def initialize_export(cls, workbook=None, columns=None, row_number=None, query_set=None, **kwargs):
        _today = date.today()
        query_params = kwargs.get('query_params')
        report_type = query_params.get('report_type', 'baseline')
        report_year = query_params.get('year', _today.year)
        scg_wise_field_values = dict()

        approved_scg_report_ids = ApprovedSCGMonthlyReport.objects.using(
            BWDatabaseRouter.get_export_database_name()
        ).distinct('parent_id').values_list(
            'pk', flat=True).order_by('parent_id', '-on_spot_creation_time')

        if report_type == 'baseline':
            scg_report_export_queryset = ApprovedSCGMonthlyReport.objects.using(
                BWDatabaseRouter.get_export_database_name()
            ).filter(
                is_baseline=True,
                pk__in=list(approved_scg_report_ids)
            ).order_by('-on_spot_creation_time')
        else:
            scg_report_export_queryset = ApprovedSCGMonthlyReport.objects.using(
                BWDatabaseRouter.get_export_database_name()
            ).filter(
                is_baseline=False,
                pk__in=list(approved_scg_report_ids)
            ).order_by('-on_spot_creation_time')
            _baseline_scg_report_queryset = ApprovedSCGMonthlyReport.objects.using(
                BWDatabaseRouter.get_export_database_name()
            ).filter(
                is_baseline=True,
                pk__in=list(approved_scg_report_ids)
            )
            _baseline_scg_reports = _baseline_scg_report_queryset.filter(
                field_values__field__assigned_code='002002009').values('scg_id', 'field_values__value')
            for _base_scg_report in _baseline_scg_reports:
                scg_wise_field_values[_base_scg_report['scg_id']] = _base_scg_report['field_values__value']

        scg_report_export_values = scg_report_export_queryset.annotate(
            scg_report_id=F('pk'),
            scg_report_month=F('month'),
            scg_report_year=F('year'),
            scg_report_remarks=F('remarks'),
            # scg_report_location = Concat(F('location__latitude', output_field=models.TextField()), Value(' '), F('location__longitude', output_field=models.TextField())),
            scg_id=F('scg_id'),
            scg_name=F('scg__name'),
            primary_group_id=F('scg__primary_group__assigned_code'),
            primary_group_name=F('scg__primary_group__name'),
            cdc_id=F('scg__primary_group__parent__assigned_code'),
            cluster_id=F('scg__primary_group__parent__parent__assigned_code'),
            cluster_name=F('scg__primary_group__parent__parent__name'),
            cdc_name=F('scg__primary_group__parent__name'),
            ward=F('scg__primary_group__parent__address__geography__name'),
            city=F('scg__primary_group__parent__address__geography__parent__name'),
            division=F('scg__primary_group__parent__address__geography__parent__name'),
            phone_number=F('created_by__phones__phone'),
            user_name=F('created_by__name')
        ).values(
            'scg_report_id', 'scg_report_month', 'scg_report_year', 'scg_report_remarks', 'scg_name',
            'primary_group_id',
            'primary_group_name', 'cdc_id', 'cdc_name', 'cluster_id', 'cluster_name', 'ward', 'city',
            'division', 'phone_number', 'user_name')

        scg_report_extra_values = scg_report_export_queryset.annotate(
            scg_report_id=F('pk'), scg_id=F('scg_id'),
            field_name=F('field_values__field__name'),
            field_value=F('field_values__value'),
            field_formula=F('field_values__field__formula'),
            field_code=F('field_values__field__assigned_code')).values(
            'scg_report_id', 'scg_id', 'field_name', 'field_value', 'field_formula', 'field_code')

        _code_wise_name_dict = dict()
        pk_wise_scg_extra_dict = dict()
        _cal_field_name_formula_dict = dict()
        _pattern = re.compile(r'@(.+?)@')

        _scg_report_field_queryset = SCGMonthlyReportField.objects.using(
            BWDatabaseRouter.get_export_database_name()
        ).values('name', 'assigned_code', 'reachout_level', 'formula')

        for _report_field in _scg_report_field_queryset:
            if _report_field['reachout_level'] == ReachoutLevelEnum.MissionControl.value:
                _cal_field_name_formula_dict[_report_field['name'].lower().strip()] = _report_field['formula']
            _code_wise_name_dict[_report_field['assigned_code']] = _report_field['name'].lower().strip()

        _scg_baseline_cal_field_name = SCGMonthlyReportField.objects.using(
            BWDatabaseRouter.get_export_database_name()
        ).filter(assigned_code='002002009').last().name.lower().strip()

        for scg_report_extra in scg_report_extra_values:
            _scg_report_id = scg_report_extra['scg_report_id']
            _scg_id = scg_report_extra['scg_id']
            _field = scg_report_extra['field_name'].lower().strip(' ')
            _value = scg_report_extra['field_value']
            if _scg_report_id not in pk_wise_scg_extra_dict:
                pk_wise_scg_extra_dict[_scg_report_id] = dict()

            if report_type == 'baseline':
                pk_wise_scg_extra_dict[_scg_report_id][_field] = _value
            elif report_type != 'baseline' and _field == _scg_baseline_cal_field_name and \
                    _scg_id in scg_wise_field_values:
                pk_wise_scg_extra_dict[_scg_report_id][_field] = scg_wise_field_values[_scg_id]
            else:
                pk_wise_scg_extra_dict[_scg_report_id][_field] = _value

        row_index = row_number + 1
        for column in columns:
            workbook.cell(row=row_index, column=column.column + 1).value = bw_titleize(column.column_name)
            workbook.cell(row=row_index, column=column.column + 1).font = Font(bold=True)
            workbook.cell(row=row_number, column=column.column + 1).alignment = Alignment(wrap_text=True)

        row_index += 1
        for scg_report_export_info in scg_report_export_values:
            _scg_report_id = scg_report_export_info['scg_report_id']
            for column in columns:
                column_index = column.column + 1
                column_name = column.column_name.lower().strip()
                if column.property_name != 'details_field':
                    _value = scg_report_export_info[column.property_name]
                else:
                    if column_name in _cal_field_name_formula_dict:
                        _formula = _cal_field_name_formula_dict[column_name]
                        _assigned_codes = _pattern.findall(_formula)
                        for _code in _assigned_codes:
                            _field_name = _code_wise_name_dict[_code] if _code in _code_wise_name_dict else None
                            _field_value = float(pk_wise_scg_extra_dict[_scg_report_id][_field_name]) \
                                if _field_name and _field_name in pk_wise_scg_extra_dict[_scg_report_id] else 0
                            _formula = _formula.replace('@' + _code + '@', str(round(_field_value, 2)))
                        try:
                            _value = eval(_formula)
                            _value = str(_value) if type(_value) != float else str(round(_value, 2))
                        except ZeroDivisionError:
                            _value = 0
                        if column_name.lower().strip() not in pk_wise_scg_extra_dict[_scg_report_id]:
                            pk_wise_scg_extra_dict[_scg_report_id][column_name.lower().strip()] = _value
                    else:
                        _value = pk_wise_scg_extra_dict[_scg_report_id][
                            column_name] if column_name in pk_wise_scg_extra_dict[_scg_report_id] else ''

                workbook.cell(row=row_index, column=column_index).value = str(_value) if _value else ''
            row_index += 1

        return workbook, row_index
