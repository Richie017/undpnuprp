import decimal
import uuid
from collections import OrderedDict
from datetime import date

import decimal
from django import forms
from django.db.models.aggregates import Sum
from django.db.models.expressions import Case, When
from django.db.models.query_utils import Q
from django.forms.forms import Form
from openpyxl.styles import NamedStyle as Style
from openpyxl.styles.alignment import Alignment

from blackwidow.core.managers.modelmanager import DomainEntityModelManager
from blackwidow.core.mixins.fieldmixin import GenericModelChoiceField
from blackwidow.core.models import Geography, ImporterConfig, ImporterColumnConfig, ErrorLog, Organization
from blackwidow.core.models.config.exporter_column_config import ExporterColumnConfig
from blackwidow.core.models.config.exporter_config import ExporterConfig
from blackwidow.engine.decorators import enable_import
from blackwidow.engine.decorators.enable_export import enable_export
from blackwidow.engine.decorators.route_partial_routes import route
from blackwidow.engine.decorators.utility import decorate, is_object_context
from blackwidow.engine.enums.modules_enum import ModuleEnum
from blackwidow.engine.enums.view_action_enum import ViewActionEnum
from blackwidow.engine.extensions.clock import Clock
from blackwidow.engine.routers.database_router import BWDatabaseRouter
from undp_nuprp.approvals.models import CityWiseMonthlyTarget

__author__ = 'Ziaul Haque'


@decorate(is_object_context, enable_export, enable_import,
          route(route='monthly-target', group='Target & Progress', module=ModuleEnum.Analysis,
                display_name='Monthly Targets', group_order=6, item_order=1))
class MonthlyTarget(CityWiseMonthlyTarget):
    objects = DomainEntityModelManager(filter={'type': 'CityWiseMonthlyTarget'})

    class Meta:
        proxy = True
        app_label = 'approvals'

    @classmethod
    def get_manage_buttons(cls):
        return [ViewActionEnum.AdvancedImport, ViewActionEnum.AdvancedExport]

    @classmethod
    def get_button_title(cls, button=ViewActionEnum.Details):
        if button == ViewActionEnum.AdvancedExport:
            return "Download Targets"
        elif button == ViewActionEnum.AdvancedImport:
            return "Update Targets"

    @classmethod
    def get_object_inline_buttons(cls):
        return []

    @classmethod
    def get_export_dependant_fields(cls):

        class AdvancedExportDependentForm(Form):
            def __init__(self, data=None, files=None, instance=None, prefix='', **kwargs):
                super(AdvancedExportDependentForm, self).__init__(data=data, files=files, prefix=prefix, **kwargs)
                today = date.today()
                year_choices = tuple()
                for y in range(2000, 2100):
                    year_choices += ((y, str(y)),)

                self.fields['city'] = \
                    GenericModelChoiceField(
                        queryset=Geography.objects.filter(level__name='Pourashava/City Corporation'),
                        label='Select City',
                        required=False,
                        widget=forms.Select(
                            attrs={
                                'class': 'select2 kpi-filter',
                                'width': '220',
                                # 'multiple': 'multiple',
                                'data-kpi-filter-role': 'city'
                            }
                        )
                    )

                self.fields['year'] = forms.ChoiceField(
                    choices=year_choices,
                    widget=forms.Select(
                        attrs={'class': 'select2', 'width': '220'}
                    ), initial=today.year
                )

        return AdvancedExportDependentForm

    @classmethod
    def header_columns(cls, month=1, year=None, quarter=None):
        yearly_target = 'Yearly Target ({0})'.format(str(year)) if year else 'Yearly Target'
        quarter = cls.get_quarter_name(quarter_id=quarter) if quarter else 'Quarter'
        columns = ['City', 'Output', 'Activities Code', 'Activities', 'Sub-activities Code', 'Sub-activities',
                   'Indicator', 'Unit', 'Year'] + [cls.get_month_name(month_id=month)] + [quarter] + [yearly_target]
        return columns

    @classmethod
    def get_target_data_rows(cls, month=None, year=None, city_ids=None, outputs=None, quarter=None, **kwargs):
        _quarter_id = int(quarter)
        queryset = MonthlyTarget.objects.filter(year=year)
        no_of_total_cities = Geography.objects.filter(level__name='Pourashava/City Corporation').values_list(
            'name', flat=True).count()
        given_cities_len = 0

        if city_ids:
            queryset = queryset.filter(city_id__in=city_ids)
            given_cities_len = len(city_ids)

        if outputs:
            q = Q(**{'output__iexact': outputs[0]})
            for o in outputs:
                q = q | Q(**{'output__iexact': o})
            queryset = queryset.filter(q)

        all_monthly_targets = queryset.values(
            'output', 'activity_code', 'activity', 'sub_activity_code',
            'sub_activity', 'indicator', 'unit', 'year'
        ).annotate(monthly_target=Sum(Case(When(month=month, then='target'))),
                   yearly_target=Sum(Case(When(month__in=list(range(1, 13)), then='target'))),
                   quarterly_target=Sum(
                       Case(When(month__in=cls.get_quarter_range(quarter_id=_quarter_id), then='target')))
                   ).order_by('activity_code', 'sub_activity_code')

        rows = []
        city_name = 'All cities' if not city_ids or no_of_total_cities == given_cities_len else '{0} Cities'.format(
            given_cities_len) if given_cities_len > 3 else ', '.join(
            list(Geography.objects.filter(pk__in=city_ids).values_list('name', flat=True)))

        for each_monthly_target in all_monthly_targets:
            row = []
            _city_name = city_name
            output = each_monthly_target['output']
            activity_code = each_monthly_target['activity_code']
            activity = each_monthly_target['activity']
            sub_activity_code = each_monthly_target['sub_activity_code']
            sub_activity = each_monthly_target['sub_activity']
            indicator = each_monthly_target['indicator']
            unit = each_monthly_target['unit']
            year = each_monthly_target['year']
            monthly_target = each_monthly_target['monthly_target'] if each_monthly_target['monthly_target'] else '-'
            yearly_target = each_monthly_target['yearly_target'] if each_monthly_target['yearly_target'] else '-'
            quarterly_target = each_monthly_target['quarterly_target'] if each_monthly_target[
                'quarterly_target'] else '-'

            row += [_city_name, output, activity_code, activity, sub_activity_code,
                    sub_activity, indicator, unit, year, monthly_target, quarterly_target, yearly_target]
            rows += [row]
        return rows

    @classmethod
    def exporter_config(cls, organization=None, **kwargs):

        ExporterConfig.objects.filter(model=cls.__name__, organization=organization).delete()
        exporter_config, created = ExporterConfig.objects.get_or_create(model=cls.__name__, organization=organization)

        columns = [
            ExporterColumnConfig(column=0, column_name='City', property_name='city_name', ignore=False),
            ExporterColumnConfig(column=1, column_name='Output', property_name='output', ignore=False),
            ExporterColumnConfig(column=2, column_name='Activities Code', property_name='activity_code', ignore=False),
            ExporterColumnConfig(column=3, column_name='Activities', property_name='activity', ignore=False),
            ExporterColumnConfig(column=4, column_name='Sub-activities Code', property_name='sub_activity_code',
                                 ignore=False),
            ExporterColumnConfig(column=5, column_name='Sub-activities', property_name='sub_activity', ignore=False),
            ExporterColumnConfig(column=6, column_name='Indicator', property_name='indicator', ignore=False),
            ExporterColumnConfig(column=7, column_name='Unit', property_name='unit', ignore=False),
            ExporterColumnConfig(column=8, column_name='Year', property_name='year', ignore=False),
        ]

        _column_index = 9
        for _month_name in cls.get_months():
            _month_index = 1
            columns.append(
                ExporterColumnConfig(column=_column_index, column_name=_month_name,
                                     property_name='month_' + str(_month_index), ignore=False)
            )
            _month_index += 1
            _column_index += 1

        for _quarter_name in cls.get_quarters():
            columns.append(ExporterColumnConfig(column=_column_index, column_name=_quarter_name,
                                                property_name=_quarter_name.replace(' ', '_').lower()))
            _column_index += 1

        columns.append(
            ExporterColumnConfig(column=_column_index, column_name='Yearly Target', property_name='yearly_target',
                                 ignore=False))

        for c in columns:
            c.save(organization=organization, **kwargs)
            exporter_config.columns.add(c)

        return exporter_config

    def export_item(self, workbook=None, columns=None, row_number=None, **kwargs):
        return self.pk, row_number + 1

    @classmethod
    def initialize_export(cls, workbook=None, columns=None, row_number=None, query_set=None, **kwargs):
        row_index = row_number + 1
        for column in columns:
            workbook.cell(row=row_index, column=column.column + 1).value = column.column_name

        group_col_start = 10
        group_col_end = 21
        center_style = Style(name='center_style', alignment=Alignment(horizontal='center', vertical='center'))

        workbook.merge_cells(
            start_row=row_index - 1, end_row=row_index - 1,
            start_column=group_col_start, end_column=group_col_end
        )
        workbook.cell(row=row_index - 1, column=group_col_start).value = 'Monthly Targets'
        workbook.cell(row=row_index - 1, column=group_col_start).style = center_style

        row_index += 1
        _today = date.today()
        query_params = kwargs.get('query_params')
        target_month = int(query_params.get('month', _today.month))
        target_year = int(query_params.get('year', _today.year))
        city_param = query_params.get('city', None)

        city_ids = None
        if city_param:
            city_ids = [int(x) for x in city_param.split(',')]

        queryset = MonthlyTarget.objects.using(BWDatabaseRouter.get_export_database_name()).all()
        if target_year:
            queryset = queryset.filter(year=target_year)
        if city_ids:
            queryset = queryset.filter(city_id__in=city_ids)

        target_dict = OrderedDict()

        monthly_target_objects = queryset.values(
            'city__name', 'year', 'month', 'output', 'activity_code', 'activity',
            'sub_activity_code', 'sub_activity', 'indicator', 'unit', 'target'
        ).order_by('city__name', 'activity_code', 'sub_activity_code', 'month')

        for _each_target in monthly_target_objects:
            city_name = _each_target['city__name']
            year = _each_target['year']
            month = _each_target['month']
            output = _each_target['output']
            activity_code = _each_target['activity_code']
            activity = _each_target['activity']
            sub_activity_code = _each_target['sub_activity_code']
            sub_activity = _each_target['sub_activity']
            indicator = _each_target['indicator']
            unit = _each_target['unit']
            target = _each_target['target'] if _each_target['target'] else 0

            _key = (city_name, output, activity_code, activity, sub_activity_code, sub_activity, indicator, unit, year)

            if _key not in target_dict.keys():
                target_dict[_key] = {
                    'city_name': city_name,
                    'output': output,
                    'activity_code': activity_code,
                    'activity': activity,
                    'sub_activity_code': sub_activity_code,
                    'sub_activity': sub_activity,
                    'indicator': indicator,
                    'unit': unit,
                    'year': year,
                    'data': []
                }

            target_dict[_key]['data'].append((month, target))

        for each_item in target_dict.values():
            column_index = 1
            for column in columns:
                column_index = column.column + 1
                _value = each_item[column.property_name]
                workbook.cell(row=row_index, column=column_index).value = str(_value) if _value else ''

                if column_index > 8:
                    break

            _data = each_item['data']
            _yearly_target = 0

            quarterly_targets = [0] * len(cls.get_quarters())

            for _d in _data:
                column_index += 1
                _month, _target = _d
                _yearly_target += _target
                _q = int((_month - 1) / 3)
                quarterly_targets[_q] += _target
                workbook.cell(row=row_index, column=column_index).value = str(_target) if _target else '-'

            for _q_target in quarterly_targets:
                column_index += 1
                workbook.cell(row=row_index, column=column_index).value = str(_q_target) if _q_target else '-'

            workbook.cell(row=row_index, column=column_index + 1).value = str(_yearly_target) if _yearly_target else '-'

            row_index += 1

        return workbook, row_index

    @classmethod
    def finalize_export(cls, workbook=None, row_count=None, **kwargs):
        file_name = '%s_%s' % (cls.__name__, Clock.timestamp())
        return workbook, file_name

    @classmethod
    def importer_config(cls, organization=None, **kwargs):
        ImporterConfig.objects.filter(model=cls.__name__, organization=organization).delete()
        importer_config, result = ImporterConfig.objects.get_or_create(model=cls.__name__, organization=organization)
        importer_config.starting_row = 2
        importer_config.starting_column = 0
        importer_config.save(**kwargs)
        columns = [
            ImporterColumnConfig(column=0, column_name='City', property_name='city_name', ignore=False),
            ImporterColumnConfig(column=1, column_name='Output', property_name='output', ignore=False),
            ImporterColumnConfig(column=2, column_name='Activities Code', property_name='activity_code', ignore=False),
            ImporterColumnConfig(column=3, column_name='Activities', property_name='activity', ignore=False),
            ImporterColumnConfig(column=4, column_name='Sub-activities Code', property_name='sub_activity_code',
                                 ignore=False),
            ImporterColumnConfig(column=5, column_name='Sub-activities', property_name='sub_activity', ignore=False),
            ImporterColumnConfig(column=6, column_name='Indicator', property_name='indicator', ignore=False),
            ImporterColumnConfig(column=7, column_name='Unit', property_name='unit', ignore=False),
            ImporterColumnConfig(column=8, column_name='Year', property_name='year', ignore=False),
        ]

        _column_index = 9
        for _month_name in cls.get_months():
            _month_index = 1
            columns.append(
                ImporterColumnConfig(column=_column_index, column_name=_month_name,
                                     property_name='month_' + str(_month_index), ignore=False)
            )
            _month_index += 1
            _column_index += 1

        for _quarter_name in cls.get_quarters():
            columns.append(ImporterColumnConfig(column=_column_index, column_name=_quarter_name,
                                                property_name=_quarter_name.replace(' ', '_').lower()))
            _column_index += 1

        columns.append(
            ImporterColumnConfig(column=_column_index, column_name='Yearly Target', property_name='yearly_target',
                                 ignore=False))

        for c in columns:
            c.save(organization=organization, **kwargs)
            importer_config.columns.add(c)
        return importer_config

    @classmethod
    def import_item(cls, config, sorted_columns, data, user=None, **kwargs):
        return data

    @classmethod
    def post_processing_import_completed(cls, items=[], user=None, organization=None, **kwargs):
        cls.process_target_import(items=items, organization=organization, **kwargs)

    @classmethod
    def process_target_import(cls, items, organization=None, **kwargs):
        """
            We keep monthly target as month = 12, and quarterly target as Q1: month = 13, Q2: month = 14,
            Q3: month = 14, Q4: month = 14 respectively
        """
        organization = organization if organization else Organization.get_organization_from_cache()
        targets = OrderedDict()
        cities = OrderedDict()
        targets_create = list()
        targets_update = list()

        for target in cls.objects.all():
            city_name = target.city.name.lower()
            output = target.output.lower()
            activity_code = target.activity_code.lower()
            activity = target.activity.lower()
            sub_activity_code = target.sub_activity_code.lower()
            sub_activity = target.sub_activity.lower()
            indicator = target.indicator.lower()
            unit = target.unit.lower()
            year = target.year
            month = target.month
            _key = (city_name, activity_code, sub_activity_code, year, month)
            targets[_key] = target

        for city in Geography.objects.filter(level__name='Pourashava/City Corporation'):
            cities[city.name.lower()] = city

        creatable_target_keys = list()
        time_now = Clock.timestamp()
        for index, item in enumerate(items):
            try:
                city_name = str(item['0'])
                output = str(item['1'])
                activity_code = str(item['2'])
                activity = str(item['3'])
                sub_activity_code = str(item['4'])
                sub_activity = str(item['5'])
                indicator = str(item['6'])
                unit = str(item['7'])
                year = int(item['8'])

                if city_name is None or len(city_name.replace(' ', '')) <= 0 \
                        or output is None or len(output.replace(' ', '')) <= 0 \
                        or activity_code is None or len(activity_code.replace(' ', '')) <= 0 \
                        or activity is None or len(activity.replace(' ', '')) <= 0 \
                        or sub_activity_code is None or len(sub_activity_code.replace(' ', '')) <= 0 \
                        or sub_activity is None or len(sub_activity.replace(' ', '')) <= 0 \
                        or indicator is None or len(indicator.replace(' ', '')) <= 0 \
                        or unit is None or len(unit.replace(' ', '')) <= 0 \
                        or year is None:
                    continue

                if city_name.lower() not in cities.keys():
                    continue

                city_id = cities[city_name.lower()].pk
                month_index = 8

                for _month_idx in list(range(1, 13)):
                    target = item[str(month_index + _month_idx)]
                    try:
                        target = decimal.Decimal(target)
                    except:
                        target = 0

                    _key = (city_name.lower(), activity_code.lower(), sub_activity_code.lower(), year, _month_idx)
                    if _key in targets.keys():
                        monthly_target = targets[_key]
                        monthly_target.target = target
                        monthly_target.tsync_id = uuid.uuid4() \
                            if monthly_target.tsync_id is None else monthly_target.tsync_id
                        monthly_target.activity = activity
                        monthly_target.sub_activity = sub_activity
                        monthly_target.indicator = indicator
                        monthly_target.unit = unit
                        monthly_target.output = output
                        monthly_target.last_updated = time_now
                        time_now += 1
                        targets_update.append(monthly_target)

                    elif _key not in creatable_target_keys:
                        monthly_target = cls(
                            organization_id=organization.pk,
                            city_id=city_id,
                            year=year,
                            month=_month_idx,
                            output=output,
                            activity_code=activity_code,
                            activity=activity,
                            sub_activity_code=sub_activity_code,
                            sub_activity=sub_activity,
                            indicator=indicator,
                            unit=unit
                        )

                        monthly_target.target = target
                        monthly_target.date_created = time_now
                        time_now += 1
                        monthly_target.type = CityWiseMonthlyTarget.__name__
                        monthly_target.tsync_id = uuid.uuid4() \
                            if monthly_target.tsync_id is None else monthly_target.tsync_id
                        monthly_target.last_updated = time_now
                        time_now += 1
                        targets_create.append(monthly_target)
                        creatable_target_keys.append(_key)

            except Exception as exp:
                ErrorLog.log(exp, organization=organization)

        if len(targets_create) > 0:
            cls.objects.bulk_create(targets_create)
        if len(targets_update) > 0:
            cls.objects.bulk_update(targets_update, batch_size=100)
