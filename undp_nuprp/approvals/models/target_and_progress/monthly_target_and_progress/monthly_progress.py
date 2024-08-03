import decimal
import uuid
from collections import OrderedDict
from datetime import date

from django import forms
from django.db.models.aggregates import Sum
from django.db.models.expressions import Case, When
from django.db.models.query_utils import Q
from openpyxl.styles import NamedStyle as Style
from openpyxl.styles.alignment import Alignment

from blackwidow.core.managers.modelmanager import DomainEntityModelManager
from blackwidow.core.mixins.fieldmixin import GenericModelChoiceField
from blackwidow.core.models import Geography, ImporterConfig, ImporterColumnConfig, Organization, ErrorLog
from blackwidow.core.models.config.exporter_column_config import ExporterColumnConfig
from blackwidow.core.models.config.exporter_config import ExporterConfig
from blackwidow.engine.decorators import enable_import
from blackwidow.engine.decorators.enable_export import enable_export
from blackwidow.engine.decorators.route_partial_routes import route
from blackwidow.engine.decorators.utility import decorate, is_object_context
from blackwidow.engine.enums.modules_enum import ModuleEnum
from blackwidow.engine.enums.view_action_enum import ViewActionEnum
from blackwidow.engine.extensions.clock import Clock
from blackwidow.engine.routers.database_router import BWDatabaseRouter
from undp_nuprp.approvals.models import CityWiseMonthlyTarget

__author__ = 'Ziaul Haque'

YearlyTargetMonth = 17


@decorate(is_object_context, enable_export, enable_import,
          route(route='monthly-progress', group='Target & Progress', module=ModuleEnum.Analysis,
                display_name='Monthly Progress', group_order=6, item_order=2))
class MonthlyProgress(CityWiseMonthlyTarget):
    objects = DomainEntityModelManager(filter={'type': 'CityWiseMonthlyTarget'})

    class Meta:
        proxy = True
        app_label = 'approvals'

    @classmethod
    def get_manage_buttons(cls):
        return [ViewActionEnum.AdvancedImport, ViewActionEnum.AdvancedExport]

    @classmethod
    def get_button_title(cls, button=ViewActionEnum.Details):
        if button == ViewActionEnum.AdvancedExport:
            return "Download Progress"
        elif button == ViewActionEnum.AdvancedImport:
            return "Update Progress"

    @classmethod
    def get_object_inline_buttons(cls):
        return []

    @classmethod
    def get_export_dependant_fields(cls):

        class AdvancedExportDependentForm(forms.Form):
            def __init__(self, data=None, files=None, instance=None, prefix='', **kwargs):
                super(AdvancedExportDependentForm, self).__init__(data=data, files=files, prefix=prefix, **kwargs)
                today = date.today()
                year_choices = tuple()
                for y in range(2000, 2100):
                    year_choices += ((y, str(y)),)

                self.fields['city'] = \
                    GenericModelChoiceField(
                        queryset=Geography.objects.filter(level__name='Pourashava/City Corporation'),
                        label='Select City',
                        required=False,
                        widget=forms.Select(
                            attrs={
                                'class': 'select2 kpi-filter',
                                'width': '220',
                                # 'multiple': 'multiple',
                                'data-kpi-filter-role': 'city'
                            }
                        )
                    )

                self.fields['year'] = forms.ChoiceField(
                    choices=year_choices,
                    widget=forms.Select(
                        attrs={'class': 'select2', 'width': '220'}
                    ), initial=today.year
                )

        return AdvancedExportDependentForm

    @classmethod
    def header_columns(cls, month=1, year=None, quarter=None):
        yearly_target = 'Yearly (Target {0})'.format(str(year)) if year else 'Yearly Target'
        quarter = cls.get_quarter_name(quarter_id=quarter) if quarter else 'Quarter'
        yearly_achievement = 'Yearly (Achievement {0})'.format(str(year)) if year else 'Yearly Achievement'
        columns = ['City', 'Output', 'Activities Code', 'Activities', 'Sub-activities Code', 'Sub-activities',
                   'Indicator', 'Unit', 'Year'] + \
                  [cls.get_month_name(month_id=month) + ' (Target)',
                   cls.get_month_name(month_id=month) + ' (Achievement)'] + \
                  [quarter + ' (Target)', quarter + ' (Achievement)'] + \
                  [yearly_target, yearly_achievement, 'Cumulative (achievement)']
        return columns

    @classmethod
    def get_progress_data_rows(cls, month=None, year=None, city_ids=None, outputs=None, quarter=None, **kwargs):
        queryset = MonthlyProgress.objects.filter(year=year)
        cum_queryset = MonthlyProgress.objects.exclude(month=YearlyTargetMonth)

        no_of_total_cities = Geography.objects.filter(level__name='Pourashava/City Corporation').values_list(
            'name', flat=True).count()
        given_cities_len = 0

        if city_ids:
            queryset = queryset.filter(city_id__in=city_ids)
            cum_queryset = cum_queryset.filter(city_id__in=city_ids)
            given_cities_len = len(city_ids)

        if outputs:
            q = Q(**{'output__iexact': outputs[0]})
            for o in outputs:
                q = q | Q(**{'output__iexact': o})

            queryset = queryset.filter(q)
            cum_queryset = cum_queryset.filter(q)

        cum_queryset = cum_queryset.values(
            'output', 'activity_code', 'activity', 'sub_activity_code', 'sub_activity', 'indicator', 'unit').annotate(
            cumulative_achievement=Sum(
                Case(When(Q(month__in=list(range(1, YearlyTargetMonth)), year__lt=year + 1),
                          then='achieved')))).order_by('activity_code', 'sub_activity_code')

        cum_ach_dict = OrderedDict()
        for _cum_ach_dict in cum_queryset:
            _key = (_cum_ach_dict['output'], _cum_ach_dict['activity_code'], _cum_ach_dict['activity'],
                    _cum_ach_dict['sub_activity_code'], _cum_ach_dict['sub_activity'], _cum_ach_dict['indicator'],
                    _cum_ach_dict['unit'])
            cum_ach_dict[_key] = _cum_ach_dict['cumulative_achievement']

        all_monthly_targets = queryset.values(
            'output', 'activity_code', 'activity', 'sub_activity_code', 'sub_activity', 'indicator', 'unit').annotate(
            monthly_target=Sum(Case(When(Q(month=month, year=year), then='target'))),
            monthly_achieved=Sum(Case(When(Q(month=month, year=year), then='achieved'))),
            yearly_target=Sum(Case(When(Q(month__in=range(1, 13), year=year), then='target'))),
            yearly_achieved=Sum(
                Case(When(Q(month__in=list(range(1, YearlyTargetMonth)), year=year), then='achieved'))),
            quarterly_target=Sum(
                Case(When(Q(month__in=cls.get_quarter_range(quarter_id=int(quarter)), year=year), then='target'))),
            quarterly_achieved=Sum(Case(When(Q(month__in=cls.get_quarter_range(quarter), year=year), then='achieved'))),
        ).order_by(
            'activity_code', 'sub_activity_code')

        rows = []
        city_name = 'All cities' if not city_ids or no_of_total_cities == given_cities_len else '{0} Cities'.format(
            given_cities_len) if given_cities_len > 3 else ', '.join(
            list(Geography.objects.filter(pk__in=city_ids).values_list('name', flat=True)))

        for each_monthly_target in all_monthly_targets:
            row = []
            output = each_monthly_target['output']
            activity_code = each_monthly_target['activity_code']
            activity = each_monthly_target['activity']
            sub_activity_code = each_monthly_target['sub_activity_code']
            sub_activity = each_monthly_target['sub_activity']
            indicator = each_monthly_target['indicator']
            unit = each_monthly_target['unit']
            monthly_target = each_monthly_target['monthly_target'] if each_monthly_target['monthly_target'] else '-'
            monthly_achieved = each_monthly_target['monthly_achieved'] if each_monthly_target[
                'monthly_achieved'] else '-'
            quarterly_target = each_monthly_target['quarterly_target'] if each_monthly_target[
                'quarterly_target'] else '-'
            quarterly_achieved = each_monthly_target['quarterly_achieved'] if each_monthly_target[
                'quarterly_achieved'] else '-'
            yearly_target = each_monthly_target['yearly_target'] if each_monthly_target['yearly_target'] else '-'
            yearly_achieved = each_monthly_target['yearly_achieved'] if each_monthly_target['yearly_achieved'] else '-'
            _key = (output, activity_code, activity, sub_activity_code, sub_activity, indicator, unit)
            cumulative_achievement = cum_ach_dict[_key] if _key in cum_ach_dict and cum_ach_dict[_key] else '-'

            row += [city_name, output, activity_code, activity, sub_activity_code, sub_activity, indicator, unit,
                    year, monthly_target, monthly_achieved, quarterly_target, quarterly_achieved, yearly_target,
                    yearly_achieved, cumulative_achievement]
            rows += [row]
        return rows

    @classmethod
    def exporter_config(cls, organization=None, **kwargs):

        ExporterConfig.objects.filter(model=cls.__name__, organization=organization).delete()
        exporter_config, created = ExporterConfig.objects.get_or_create(model=cls.__name__, organization=organization)

        columns = [
            ExporterColumnConfig(column=0, column_name='City', property_name='city_name', ignore=False),
            ExporterColumnConfig(column=1, column_name='Output', property_name='output', ignore=False),
            ExporterColumnConfig(column=2, column_name='Activities Code', property_name='activity_code', ignore=False),
            ExporterColumnConfig(column=3, column_name='Activities', property_name='activity', ignore=False),
            ExporterColumnConfig(column=4, column_name='Sub-activities Code', property_name='sub_activity_code',
                                 ignore=False),
            ExporterColumnConfig(column=5, column_name='Sub-activities', property_name='sub_activity', ignore=False),
            ExporterColumnConfig(column=6, column_name='Indicator', property_name='indicator', ignore=False),
            ExporterColumnConfig(column=7, column_name='Unit', property_name='unit', ignore=False),
            ExporterColumnConfig(column=8, column_name='Year', property_name='year', ignore=False),
        ]

        _column_index = 9
        for i in range(17):
            _month_index = 1
            columns.append(
                ExporterColumnConfig(column=_column_index, column_name='Target',
                                     property_name='target_' + str(_month_index), ignore=False)
            )
            _column_index += 1
            columns.append(
                ExporterColumnConfig(column=_column_index, column_name='Achievement',
                                     property_name='achievement_' + str(_month_index), ignore=False)
            )
            _month_index += 1
            _column_index += 1

        columns.append(
            ExporterColumnConfig(column=_column_index, column_name='Cumulative Achievement',
                                 property_name='cumulative_achievement', ignore=False)
        )

        for c in columns:
            c.save(organization=organization, **kwargs)
            exporter_config.columns.add(c)

        return exporter_config

    def export_item(self, workbook=None, columns=None, row_number=None, **kwargs):
        return self.pk, row_number + 1

    @classmethod
    def initialize_export(cls, workbook=None, columns=None, row_number=None, query_set=None, **kwargs):
        row_index = row_number + 1
        _today = date.today()
        query_params = kwargs.get('query_params')
        target_month = int(query_params.get('month', _today.month))
        target_year = int(query_params.get('year', _today.year))
        city_param = query_params.get('city', None)
        quarter_param = query_params.get('quarter', None)

        for column in columns:
            workbook.cell(row=row_index, column=column.column + 1).value = column.column_name

        center_style = Style(name='center_style', alignment=Alignment(horizontal='center', vertical='center'))
        col_index = 10
        for _month_name in cls.get_months() + cls.get_quarters() + ['Yearly Target ({0})'.format(str(target_year))]:
            group_col_start = col_index
            col_index += 2
            group_col_end = col_index - 1
            workbook.merge_cells(
                start_row=row_index - 1, end_row=row_index - 1,
                start_column=group_col_start, end_column=group_col_end
            )

            workbook.cell(row=row_index - 1, column=group_col_start).value = _month_name
            workbook.cell(row=row_index - 1, column=group_col_start).style = center_style

        row_index += 1
        city_ids = None
        if city_param:
            city_ids = [int(x) for x in city_param.split(',')]

        queryset = MonthlyProgress.objects.using(BWDatabaseRouter.get_export_database_name()).all()
        cum_ach_queryset = MonthlyProgress.objects.exclude(month=13)

        if target_year:
            queryset = queryset.filter(year=target_year)

        if city_ids:
            queryset = queryset.filter(city_id__in=city_ids)
            cum_ach_queryset = cum_ach_queryset.filter(city_id__in=city_ids)

        target_dict = OrderedDict()

        monthly_target_objects = queryset.values(
            'city__name', 'year', 'month', 'output', 'activity_code', 'activity',
            'sub_activity_code', 'sub_activity', 'indicator', 'unit', 'target', 'achieved'
        ).order_by('city__name', 'activity_code', 'sub_activity_code')

        cum_ach_objects = cum_ach_queryset.values(
            'city__name', 'output', 'activity_code', 'activity', 'sub_activity_code', 'sub_activity', 'indicator',
            'unit').annotate(
            cumulative_achievement=Sum(
                Case(When(Q(month__in=list(range(1, 13)), year__lt=target_year + 1), then='achieved')))
        ).order_by('city__name', 'activity_code', 'sub_activity_code')

        cum_ach_dict = OrderedDict()

        for _cum_ach_dict in cum_ach_objects:
            city_name = _cum_ach_dict['city__name']
            output = _cum_ach_dict['output']
            activity_code = _cum_ach_dict['activity_code']
            activity = _cum_ach_dict['activity']
            sub_activity_code = _cum_ach_dict['sub_activity_code']
            sub_activity = _cum_ach_dict['sub_activity']
            indicator = _cum_ach_dict['indicator']
            unit = _cum_ach_dict['unit']
            _key = (
                city_name, output, activity_code, activity, sub_activity_code, sub_activity, indicator, unit,
                target_year)
            cum_ach_dict[_key] = _cum_ach_dict['cumulative_achievement']

        for _each_target in monthly_target_objects:
            city_name = _each_target['city__name']
            year = _each_target['year']
            month = _each_target['month']
            output = _each_target['output']
            activity_code = _each_target['activity_code']
            activity = _each_target['activity']
            sub_activity_code = _each_target['sub_activity_code']
            sub_activity = _each_target['sub_activity']
            indicator = _each_target['indicator']
            unit = _each_target['unit']
            target = _each_target['target']
            achieved = _each_target['achieved']

            _key = (city_name, output, activity_code, activity, sub_activity_code, sub_activity, indicator, unit, year)

            if _key not in target_dict.keys():
                target_dict[_key] = {
                    'city_name': city_name,
                    'output': output,
                    'activity_code': activity_code,
                    'activity': activity,
                    'sub_activity_code': sub_activity_code,
                    'sub_activity': sub_activity,
                    'indicator': indicator,
                    'unit': unit,
                    'year': year,
                    'data': [(i, None, None) for i in range(1, 13)]
                }
            _data_index = month - 1
            target_dict[_key]['data'][_data_index] = (month, target, achieved)

        for _key, each_item in target_dict.items():
            column_index = 0
            for column in columns:
                column_index = column.column + 1
                if column_index > 9:
                    break
                _value = each_item[column.property_name]
                workbook.cell(row=row_index, column=column_index).value = str(_value) if _value else ''

            _data = each_item['data']
            yearly_target = 0
            yearly_achievement = 0
            month_wise_target_achieved = dict()
            quarterly_target_achievements = [(0, 0)] * len(cls.get_quarters())

            for _d in _data:
                _month, _target, _achieved = _d
                yearly_target += _target if _target else 0
                yearly_achievement += _achieved if _achieved else 0
                _q = int((_month - 1) / 3)
                _q_target, _q_achieved = quarterly_target_achievements[_q]
                _q_target += _target if _target else 0
                _q_achieved += _achieved if _achieved else 0
                quarterly_target_achievements[_q] = (_q_target, _q_achieved)

                cls.write_value_excel(workbook=workbook, row_number=row_index, column_number=column_index,
                                      value=str(_target) if _target else '-')
                cls.write_value_excel(workbook=workbook, row_number=row_index, column_number=column_index + 1,
                                      value=str(_achieved) if _achieved else '-')
                column_index += 2

            for _q_target_achievement in quarterly_target_achievements:
                _target, _achieved = _q_target_achievement
                cls.write_value_excel(workbook=workbook, row_number=row_index, column_number=column_index,
                                      value=str(_target) if _target else '-')
                cls.write_value_excel(workbook=workbook, row_number=row_index, column_number=column_index + 1,
                                      value=str(_achieved) if _achieved else '-')
                column_index += 2

            cum_achieved = str(cum_ach_dict[_key]) if _key in cum_ach_dict and cum_ach_dict[_key] else '-'

            cls.write_value_excel(workbook=workbook, row_number=row_index, column_number=column_index,
                                  value=str(yearly_target) if yearly_target else '-')
            cls.write_value_excel(workbook=workbook, row_number=row_index, column_number=column_index + 1,
                                  value=str(yearly_achievement) if yearly_achievement else '-')
            cls.write_value_excel(workbook=workbook, row_number=row_index, column_number=column_index + 2,
                                  value=cum_achieved)

            row_index += 1

        return workbook, row_index

    @classmethod
    def finalize_export(cls, workbook=None, row_count=None, **kwargs):
        file_name = '%s_%s' % (cls.__name__, Clock.timestamp())
        return workbook, file_name

    @classmethod
    def importer_config(cls, organization=None, **kwargs):
        ImporterConfig.objects.filter(model=cls.__name__, organization=organization).delete()
        importer_config, result = ImporterConfig.objects.get_or_create(model=cls.__name__, organization=organization)
        importer_config.starting_row = 2
        importer_config.starting_column = 0
        importer_config.save(**kwargs)
        columns = [
            ImporterColumnConfig(column=0, column_name='City', property_name='city_name', ignore=False),
            ImporterColumnConfig(column=1, column_name='Output', property_name='output', ignore=False),
            ImporterColumnConfig(column=2, column_name='Activities Code', property_name='activity_code', ignore=False),
            ImporterColumnConfig(column=3, column_name='Activities', property_name='activity', ignore=False),
            ImporterColumnConfig(column=4, column_name='Sub-activities Code', property_name='sub_activity_code',
                                 ignore=False),
            ImporterColumnConfig(column=5, column_name='Sub-activities', property_name='sub_activity', ignore=False),
            ImporterColumnConfig(column=6, column_name='Indicator', property_name='indicator', ignore=False),
            ImporterColumnConfig(column=7, column_name='Unit', property_name='unit', ignore=False),
            ImporterColumnConfig(column=8, column_name='Year', property_name='year', ignore=False),
        ]

        _column_index = 9
        for i in range(17):
            _month_index = 1
            columns.append(
                ImporterColumnConfig(column=_column_index, column_name='Target',
                                     property_name='target_' + str(_month_index), ignore=False)
            )
            _column_index += 1
            columns.append(
                ImporterColumnConfig(column=_column_index, column_name='Achievement',
                                     property_name='achievement_' + str(_month_index), ignore=False)
            )
            _month_index += 1
            _column_index += 1

        for c in columns:
            c.save(organization=organization, **kwargs)
            importer_config.columns.add(c)
        return importer_config

    @classmethod
    def import_item(cls, config, sorted_columns, data, user=None, **kwargs):
        return data

    @classmethod
    def post_processing_import_completed(cls, items=[], user=None, organization=None, **kwargs):
        cls.process_progress_import(items=items, organization=organization, **kwargs)

    @classmethod
    def process_progress_import(cls, items, organization=None, **kwargs):
        """
            We keep monthly progress as month = 12, and quarterly progress as Q1: month = 13, Q2: month = 14,
            Q3: month = 14, Q4: month = 14 respectively
        """
        organization = organization if organization else Organization.get_organization_from_cache()
        progresses = OrderedDict()
        progresses_update = list()

        for progress in cls.objects.all():
            city_name = progress.city.name.lower()
            output = progress.output.lower()
            activity_code = progress.activity_code.lower()
            activity = progress.activity.lower()
            sub_activity_code = progress.sub_activity_code.lower()
            sub_activity = progress.sub_activity.lower()
            indicator = progress.indicator.lower()
            unit = progress.unit.lower()
            year = progress.year
            month = progress.month
            _key = (city_name, output, activity_code, activity, sub_activity_code,
                    sub_activity, indicator, unit, year, month)
            progresses[_key] = progress

        time_now = Clock.timestamp()
        for index, item in enumerate(items):
            try:
                city_name = str(item['0'])
                output = str(item['1'])
                activity_code = str(item['2'])
                activity = str(item['3'])
                sub_activity_code = str(item['4'])
                sub_activity = str(item['5'])
                indicator = str(item['6'])
                unit = str(item['7'])
                year = int(item['8'])

                if city_name is None or len(city_name.replace(' ', '')) <= 0 \
                        or output is None or len(output.replace(' ', '')) <= 0 \
                        or activity_code is None or len(activity_code.replace(' ', '')) <= 0 \
                        or activity is None or len(activity.replace(' ', '')) <= 0 \
                        or sub_activity_code is None or len(sub_activity_code.replace(' ', '')) <= 0 \
                        or sub_activity is None or len(sub_activity.replace(' ', '')) <= 0 \
                        or indicator is None or len(indicator.replace(' ', '')) <= 0 \
                        or unit is None or len(unit.replace(' ', '')) <= 0 \
                        or year is None:
                    continue

                column_index = 8
                _month_index = 0
                for _month_name in cls.get_months():
                    month = cls.map_month_id(_month_name) if _month_name != '' else 13
                    target = item[str(column_index + month + _month_index)]
                    _month_index += 1
                    achieved = item[str(column_index + month + _month_index)]
                    try:
                        achieved = decimal.Decimal(achieved)
                    except:
                        achieved = 0

                    _key = (city_name.lower(), output.lower(), activity_code.lower(), activity.lower(),
                            sub_activity_code.lower(), sub_activity.lower(), indicator.lower(), unit.lower(), year,
                            month)

                    if _key in progresses.keys():
                        monthly_progress = progresses[_key]
                        monthly_progress.achieved = achieved
                        monthly_progress.tsync_id = uuid.uuid4() \
                            if monthly_progress.tsync_id is None else monthly_progress.tsync_id
                        monthly_progress.last_updated = time_now
                        time_now += 1
                        progresses_update.append(monthly_progress)

            except Exception as exp:
                ErrorLog.log(exp, organization=organization)

        if len(progresses_update) > 0:
            cls.objects.bulk_update(progresses_update, batch_size=100)
