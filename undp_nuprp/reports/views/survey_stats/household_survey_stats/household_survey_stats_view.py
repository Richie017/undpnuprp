"""
Created by tareq on 4/25/17
"""
import os

from crequest.middleware import CrequestMiddleware
from django import forms
from django.conf import settings
from django.contrib import messages
from django.contrib.postgres.forms.ranges import DateRangeField
from django.db.models.aggregates import Min, Max
from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fonts import Font

from blackwidow.core.mixins.fieldmixin.multiple_choice_field_mixin import GenericModelChoiceField
from blackwidow.core.models.file.exportfileobject import ExportFileObject
from blackwidow.core.models.geography.geography import Geography
from blackwidow.core.models.geography.geography_level import GeographyLevel
from blackwidow.core.models.log.error_log import ErrorLog
from blackwidow.engine.decorators.utility import decorate, get_models_with_decorator
from blackwidow.engine.decorators.view_decorators import override_view
from blackwidow.engine.enums.view_action_enum import ViewActionEnum
from blackwidow.engine.extensions.clock import Clock
from blackwidow.engine.file_handlers.aws_file_writer import AWSFileWriter
from blackwidow.engine.routers.database_router import BWDatabaseRouter
from config.aws_s3_config import MEDIA_DIRECTORY
from config.model_json_cache import MODEL_JASON_URL
from undp_nuprp.reports.models.survey_stats.household_survey_stats.household_survey_stats import HouseholdSurveyStats
from undp_nuprp.reports.views.base.base_report import GenericReportView
from undp_nuprp.survey.models import Survey
from undp_nuprp.survey.models.response.survey_response import SurveyResponse

S3_STATIC_ENABLED = settings.S3_STATIC_ENABLED
EXPORT_FILE_ROOT = settings.EXPORT_FILE_ROOT
INSTALLED_APPS = settings.INSTALLED_APPS

__author__ = 'Tareq'


@decorate(override_view(model=HouseholdSurveyStats, view=ViewActionEnum.Manage))
class HouseholdSurveyStatsView(GenericReportView):
    def get(self, request, *args, **kwargs):
        if request.GET.get('export', 'False') == 'True':
            survey_ids, city_ids, ward_ids, selected_wards, f_date, t_date = self.process_parameters()

            dest_filename = self.__class__.generate_excel(
                request, surveys=survey_ids,
                cities=city_ids,
                wards=ward_ids,
                domain=selected_wards,
                time_from=f_date,
                time_to=t_date
            )
            if dest_filename is not None:
                if self.is_json_request(request):
                    return super(HouseholdSurveyStatsView, self).get_json_response(
                        self.convert_context_to_json({
                            'success': True,
                            "url": '/export-files/download/' + str(dest_filename.pk)
                        }), **kwargs
                    )
                _message = "Enumerator Survey Stats has been successfully exported to files. " \
                           "<ol><li>" + dest_filename.file.name + "</li>" + \
                           "Please visit the <strong>Exported Files</strong> section to download them." \
                           "<a class='btn btn-danger btn-small' href='/export-files/' %}'>View Exported Files</a>"

                messages.success(request, _message, "permanent")

        return super(HouseholdSurveyStatsView, self).get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(HouseholdSurveyStatsView, self).get_context_data(**kwargs)
        context['title'] = "Survey Stats by Location"
        return context

    def get_report_parameters(self, **kwargs):
        parameters = super(HouseholdSurveyStatsView, self).get_report_parameters(**kwargs)
        city_level = GeographyLevel.objects.using(
            BWDatabaseRouter.get_read_database_name()
        ).filter(name__icontains='Pourashava/City Corporation').first()

        request = CrequestMiddleware.get_request()
        user = request.c_user
        user_role = user.role.name

        filter_roles = get_models_with_decorator(
            decorator_name='has_data_filter',
            apps=INSTALLED_APPS,
            include_class=False
        )
        if user_role in filter_roles:
            json_suffix = '_' + str(user.pk)
        else:
            json_suffix = ''

        parameters['G1'] = self.get_wrapped_parameters((
            {
                'name': 'survey',
                'field': GenericModelChoiceField(
                    queryset=Survey.objects.using(BWDatabaseRouter.get_read_database_name()).all(),
                    label='Survey Type',
                    empty_label=None,
                    required=False,
                    widget=forms.Select(
                        attrs={
                            'class': 'select2',
                            'width': '220',
                            'multiple': 'multiple'
                        }
                    )
                )
            },
        ))
        parameters['G2'] = self.get_wrapped_parameters((
            {
                'name': 'city_corporation',
                'field': GenericModelChoiceField(
                    queryset=Geography.objects.using(BWDatabaseRouter.get_read_database_name()).filter(
                        level_id=city_level),
                    label='City/Town',
                    required=False,
                    empty_label=None,
                    widget=forms.Select(
                        attrs={
                            'width': '220',
                            'multiple': 'multiple',
                            'class': 'select2',
                            'data-child': 'ward'
                        }
                    )
                )
            },
        ))
        parameters['G3'] = self.get_wrapped_parameters((
            {
                'name': 'ward',
                'field': forms.CharField(
                    label='Ward',
                    required=False,
                    widget=forms.TextInput(
                        attrs={
                            'width': '220',
                            'multiple': 'multiple',
                            'class': 'bw-select2',
                            'data-load-none': 'true',
                            'data-depends-on': 'city_corporation',
                            'data-depends-property': 'parent',
                            'data-js-url': '{0}{1}{2}.js'.format(
                                MODEL_JASON_URL, Geography.__name__.lower(), json_suffix)
                        }
                    )
                )
            },
        ))
        parameters['G4'] = self.get_wrapped_parameters((
            {
                'name': 'date_range',
                'field': DateRangeField(
                    widget=forms.TextInput(
                        attrs={
                            'data-initial-empty': "true",
                            'class': 'date-range-picker'
                        }
                    )
                )
            },
        ))
        return parameters

    def get_template_names(self):
        return ['reports/survey-stats/survey-stats.html']

    def process_parameters(self):
        survey = self.extract_parameter('survey')
        date_range = self.extract_parameter('date_range')
        city_corporation = self.extract_parameter('city_corporation')
        ward = self.extract_parameter('ward')

        survey_ids = None
        city_ids = None
        ward_ids = None
        selected_wards = None
        if survey is not None:
            survey_ids = [int(id) for id in survey.split(',')]
        if city_corporation is not None:
            city_ids = [int(id) for id in city_corporation.split(',')]
            selected_wards = Geography.objects.using(BWDatabaseRouter.get_read_database_name()).filter(
                level__name__icontains='ward').filter(parent_id__in=city_ids)
        if ward is not None:
            ward_ids = [int(id) for id in ward.split(',')]
            selected_wards = Geography.objects.using(BWDatabaseRouter.get_read_database_name()).filter(
                level__name__icontains='ward').filter(pk__in=ward_ids)

        f_date, t_date = Clock.date_range_from_str(date_range)
        if f_date is None:
            f_date = SurveyResponse.objects.using(
                BWDatabaseRouter.get_read_database_name()
            ).aggregate(Min('survey_time'))['survey_time__min']
        if t_date is None:
            t_date = SurveyResponse.objects.using(
                BWDatabaseRouter.get_read_database_name()
            ).aggregate(max_time=Max('survey_time'))['max_time']

        if selected_wards is not None:
            selected_wards = selected_wards.values_list('pk', flat=True)

        return survey_ids, city_ids, ward_ids, selected_wards, f_date, t_date

    def get_json_response(self, content, **kwargs):
        survey_ids, city_ids, ward_ids, selected_wards, f_date, t_date = self.process_parameters()

        report, fixed_columns = HouseholdSurveyStats.build_report(
            surveys=survey_ids, cities=city_ids,
            wards=ward_ids, domain=selected_wards,
            time_from=f_date, time_to=t_date
        )

        data_dict = dict()
        data_dict['report'] = report
        data_dict['fixed_columns'] = fixed_columns
        return super(HouseholdSurveyStatsView, self).get_json_response(
            self.convert_context_to_json(data_dict), **kwargs
        )

    @classmethod
    def generate_excel(cls, request, surveys=None, cities=None, wards=None, domain=None, time_from=None,
                       time_to=None):
        try:
            path = os.path.join(EXPORT_FILE_ROOT)
            if not os.path.exists(path):
                os.makedirs(path)
            wb = Workbook()
            dest_filename = 'Survey_Stats_' + str(Clock.utcnow().strftime('%d%m%Y%H%M'))
            file_path = path + os.sep + dest_filename + '.xlsx'
            ws = wb.create_sheet(index=0)
            row_number = 1
            column_number = 1
            ws.title = 'Survey Stats'
            title_cell_style = ws.cell(row=row_number, column=column_number)
            title_cell_style.value = "Survey Stats"
            # row = ws.row_dimensions[1]
            title_cell_style.font = Font(
                name='Calibri',
                size=10,
                bold=True,
                vertAlign=None)
            title_cell_style.alignment = Alignment(horizontal='general')
            ws.merge_cells(start_row=row_number, start_column=column_number, end_row=row_number,
                           end_column=column_number + 8)

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            row_number = 1

            report, fixed_columns = HouseholdSurveyStats.build_report(
                surveys=surveys, cities=cities,
                wards=wards, domain=domain,
                time_from=time_from, time_to=time_to
            )

            row_number = row_number + 2

            report_header = report[0]
            report_footer = report[-1]
            report_body = report[1:-1]
            col = 1
            ws.row_dimensions[row_number].height = 45
            for _col in report_header:
                ws.cell(row=row_number, column=col).value = _col
                ws.cell(row=row_number, column=col).font = Font(bold=True)
                ws.cell(row=row_number, column=col).border = thin_border
                ws.cell(row=row_number, column=col).alignment = Alignment(wrap_text=True)
                col += 1
            row_number += 1
            for _row in report_body:
                col = 1
                for _col in _row:
                    ws.cell(row=row_number, column=col).value = _col
                    ws.cell(row=row_number, column=col).border = thin_border
                    col += 1
                row_number += 1
            col = 1
            for _col in report_footer:
                ws.cell(row=row_number, column=col).value = _col
                ws.cell(row=row_number, column=col).font = Font(bold=True)
                ws.cell(row=row_number, column=col).border = thin_border
                col += 1
            row_number += 1

            wb.save(filename=file_path)

            # Uploading the exported file to AMAZON S3
            if S3_STATIC_ENABLED:
                from config.aws_s3_config import STATIC_EXPORT_MEDIA_DIRECTORY
                try:
                    with open(file_path, 'rb') as content_file:
                        content = content_file.read()
                        s3_file_name = STATIC_EXPORT_MEDIA_DIRECTORY + "/" + str(dest_filename) + '.xlsx'
                        file_path = s3_file_name.replace(MEDIA_DIRECTORY, '', 1)
                        AWSFileWriter.upload_file_with_content(file_name=s3_file_name, content=content)

                    # after successfully upload to AWS S3, remove local file
                    os.remove(file_path)
                except Exception as exp:
                    ErrorLog.log(exp=exp)

            export_file_object = ExportFileObject()
            export_file_object.path = file_path
            export_file_object.name = dest_filename
            export_file_object.file = file_path
            export_file_object.extension = '.xlsx'
            export_file_object.save()
            return export_file_object
        except Exception as exp:
            ErrorLog.log(exp)
        return None
