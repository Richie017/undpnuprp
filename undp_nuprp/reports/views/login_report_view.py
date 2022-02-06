import os

from django import forms
from django.conf import settings
from django.contrib import messages
from django.core.urlresolvers import reverse
from openpyxl import Workbook
from openpyxl.styles import NamedStyle as Style
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font

from blackwidow.core.mixins.fieldmixin.date_range_widget import DateRangeField
from blackwidow.core.mixins.fieldmixin.multiple_choice_field_mixin import GenericModelChoiceField
from blackwidow.core.models import ConsoleUser, Organization
from blackwidow.core.models import Role
from blackwidow.core.models.file.exportfileobject import ExportFileObject
from blackwidow.core.models.log.error_log import ErrorLog
from blackwidow.engine.decorators.utility import decorate
from blackwidow.engine.decorators.view_decorators import override_view
from blackwidow.engine.enums.view_action_enum import ViewActionEnum
from blackwidow.engine.extensions.clock import Clock
from blackwidow.engine.file_handlers.aws_file_writer import AWSFileWriter
from config.aws_s3_config import MEDIA_DIRECTORY
from undp_nuprp.reports.models import UserLoginReport
from undp_nuprp.reports.views.base.base_report import GenericReportView

EXPORT_FILE_ROOT = settings.EXPORT_FILE_ROOT

__author__ = 'Ziaul Haque'


@decorate(override_view(model=UserLoginReport, view=ViewActionEnum.Manage))
class UserLoginReportView(GenericReportView):
    def post(self, request, *args, **kwargs):
        if request.GET.get('export', 'False') == 'True':
            date_range = self.extract_parameter('date_range')
            from_date, to_date, f_date, t_date = Clock.date_range_all_from_str_no_timestamp(date_range)

            try:
                role = self.extract_parameter('role')
                users = self.extract_parameter('console_users')
                if users:
                    console_users = [int(x) for x in users.split(',')]
                else:
                    console_users = None
            except:
                role = None
                console_users = None

            dest_filename = self.__class__.generate_excel(
                request, f_date, t_date, from_date,
                to_date, role, console_users
            )
            if dest_filename is not None:
                if self.is_json_request(request):
                    return super().get_json_response(self.convert_context_to_json({
                        'success': True,
                        "url": '/export-files/download/' + str(dest_filename.pk)
                    }), **kwargs)
                messages.success(
                    request,
                    "Login Report has been successfully exported to files. "
                    "<ol><li>" + dest_filename.file.name + "</li>" +
                    "Please visit the <strong>Exported Files</strong> section to download them."
                    "<a class='btn btn-danger btn-small' href='/export-files/' %}'>View Exported Files</a>",
                    "permanent"
                )

        return super(UserLoginReportView, self).get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(UserLoginReportView, self).get_context_data(**kwargs)
        context['title'] = "Login Report"
        return context

    def get_report_parameters(self, **kwargs):
        parameters = super(UserLoginReportView, self).get_report_parameters(**kwargs)
        parameters['G2'] = self.get_wrapped_parameters((
            {
                'name': 'date_range',
                'field': DateRangeField(
                    widget=forms.TextInput(
                        attrs={
                            'data-open-direction': 'left',
                            'data-start-limit': 10,
                            'data-end-limit': 0,
                            'class': 'date-range-picker'
                        }
                    )
                )
            },
        ))
        parameters['G1'] = self.get_wrapped_parameters((
            {
                'name': 'role',
                'field': GenericModelChoiceField(
                    empty_label="All",
                    queryset=Role.objects.all(),
                    widget=forms.Select(
                        attrs={
                            'class': 'select2',
                            'width': '220'
                        }
                    )
                )
            },
            {
                'name': 'console_users',
                'field': GenericModelChoiceField(
                    label="Users",
                    queryset=ConsoleUser.objects.all(),
                    widget=forms.TextInput(
                        attrs={
                            'class': 'select2-input',
                            'width': '220',
                            'multiple': 'multiple',
                            'data-depends-on': 'role',
                            'data-depends-property': 'role:id',
                            'data-url': reverse(ConsoleUser.get_route_name(
                                action=ViewActionEnum.Manage)) + '?format=json&search=1'
                        }
                    )
                )
            },
        ))

        parameters['G3'] = self.get_wrapped_parameters(())
        parameters['G4'] = self.get_wrapped_parameters(())

        return parameters

    def get_template_names(self):
        return ['reports/login-report.html']

    def get_json_response(self, content, **kwargs):
        try:
            role = self.extract_parameter('role')
            users = self.extract_parameter('console_users')
            if users:
                console_users = [int(x) for x in users.split(',')]
            else:
                console_users = None
        except:
            role = None
            console_users = None

        date_range = self.extract_parameter('date_range')
        f_date, t_date = Clock.date_range_from_str_no_timestamp(date_range)

        data_dict = dict()
        report = UserLoginReport.build_report(
            role_filter=role, console_users=console_users, time_from=f_date, time_to=t_date, styled=True
        )
        data_dict['report'] = report
        return super(UserLoginReportView, self).get_json_response(self.convert_context_to_json(data_dict), **kwargs)

    @classmethod
    def generate_excel(cls, request, f_date, t_date, from_detail, to_detail, role=None, users=None):
        try:
            path = os.path.join(EXPORT_FILE_ROOT)
            if not os.path.exists(path):
                os.makedirs(path)
            wb = Workbook()
            dest_filename = 'Users_login_summary_' + str(Clock.utcnow().strftime('%d%m%Y%H%M'))
            file_path = path + os.sep + dest_filename + '.xlsx'
            ws = wb.create_sheet(index=0)
            row_number = 1
            column_number = 1
            ws.title = 'Login Report'
            header_cell = ws.cell(row=row_number, column=column_number)
            header_cell.value = "User Login Report between (" + str(from_detail) + "-" + str(to_detail) + ")"
            header_cell.style = Style(font=Font(
                name='Calibri',
                size=10,
                bold=True,
                vertAlign=None), alignment=Alignment(horizontal='general'))
            ws.merge_cells(
                start_row=row_number,
                start_column=column_number,
                end_row=row_number,
                end_column=column_number + 4
            )
            row_number = 3

            report = UserLoginReport.build_report(
                time_from=f_date, time_to=t_date,
                role_filter=role, console_users=users
            )
            for _row in report:
                col = 1
                for _col in _row:
                    ws.cell(row=row_number, column=col).value = _col
                    col += 1

                row_number += 1
            wb.save(filename=file_path)

            # Uploading the exported file to AMAZON S3
            if settings.S3_STATIC_ENABLED:
                from config.aws_s3_config import STATIC_EXPORT_MEDIA_DIRECTORY
                try:
                    with open(file_path, 'rb') as content_file:
                        content = content_file.read()
                        s3_file_name = STATIC_EXPORT_MEDIA_DIRECTORY + "/" + str(dest_filename) + '.xlsx'
                        file_path = s3_file_name.replace(MEDIA_DIRECTORY, '', 1)
                        AWSFileWriter.upload_file_with_content(file_name=s3_file_name, content=content)

                    # after successfully upload to AWS S3, remove local file
                    os.remove(file_path)
                except Exception as exp:
                    ErrorLog.log(exp=exp)

            export_file_object = ExportFileObject()
            export_file_object.path = file_path
            export_file_object.name = dest_filename
            export_file_object.file = file_path
            export_file_object.extension = '.xlsx'
            export_file_object.organization = Organization.objects.first()
            export_file_object.save()
            return export_file_object
        except Exception as exp:
            ErrorLog.log(exp)
        return None
